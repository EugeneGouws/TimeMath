"""Phase 0 (teacher half) -- read TT<year>.xlsx qualification tokens.

Only columns A..G are read: Teacherid | Teacher Code | sua | sub | suc | sud |
sue. Everything after `sue` (TSurname onward, and the A1..H7/P1/P2 grid) is
2026 *actual assignment* data left over from the old source, not
qualifications, and is deliberately never loaded -- it also keeps real names
out of the repo (CLAUDE.md sharing policy).

Token semantics (CLAUDE.md "Teacher file format"):
  bare code   -> grades 8..12        "SC"
  _S suffix   -> grades 10..12       "SC_S"
  _J suffix   -> grades 8..9         "SC_J"
  _<n> suffix -> that grade only     "SC_11"
Repeated tokens for the same subject UNION their grade sets. The result is
intersected with the roster-derived existence table, so a teacher is never
"qualified" for a subject in a grade where the subject does not run.

A token is a CAPABILITY, not a count -- repeats no longer mean "teaches n
classes" as they did in the legacy file.
"""
from dataclasses import dataclass
from openpyxl import load_workbook
import constants as c

ALL_GRADES = frozenset({8, 9, 10, 11, 12})
SENIOR_GRADES = frozenset({10, 11, 12})
JUNIOR_GRADES = frozenset({8, 9})


@dataclass(frozen=True)
class Teacher:
    tid: int
    code: str                       # NOT unique -- ids 1 and 3 are both "AA"
    tokens: tuple[str, ...]         # raw, as written in the sheet


def parse_teachers(path: str) -> list[Teacher]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    id_col = headers.index("Teacherid")
    code_col = headers.index("Teacher Code")

    # read su* columns dynamically -- do not hardcode C..G
    su_cols = []
    for i in range(len(headers)):
        h = headers[i]
        if isinstance(h, str) and h.strip().lower().startswith("su"):
            su_cols.append(i)
    if not su_cols:
        raise ValueError(f"{path}: no su* qualification columns found")

    teachers = []
    for row in ws.iter_rows(min_row=2):
        if row[id_col].value is None:
            continue
        tid = int(row[id_col].value)
        code = str(row[code_col].value).strip().upper()

        tokens = []
        for i in su_cols:
            v = row[i].value
            if v is None:
                continue
            tokens.append(str(v).strip().upper())

        # a completely full row suggests the source truncated a further
        # qualification -- warn, do not fail (4 such rows in Teachers2026)
        if len(tokens) == len(su_cols):
            print(f"[tt_loader] WARNING teacher {tid} ({code}) fills all "
                  f"{len(su_cols)} qualification slots -- source may have truncated")

        teachers.append(Teacher(tid=tid, code=code, tokens=tuple(tokens)))

    ids = [t.tid for t in teachers]
    if len(set(ids)) != len(ids):
        raise ValueError(f"{path}: duplicate Teacherid")
    return teachers


def parse_teacher_details(path: str) -> dict[int, dict]:
    """Read the DISPLAY columns (TSurname, TTitle, TInitials, TClassroom).

    Kept deliberately separate from parse_teachers, which reads A..G only.
    These are real people's names: the solver never needs them and must never
    see them. They exist solely so the V3 JSON export can build its
    CODE_SURNAME_GG lesson labels, and anything using this function must write
    its output under data/ (git-ignored), never into the repo.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]

    def col(name):
        return headers.index(name) if name in headers else None

    id_col = headers.index("Teacherid")
    cols = {k: col(k) for k in ("TSurname", "TTitle", "TInitials", "TClassroom")}

    details = {}
    for row in ws.iter_rows(min_row=2):
        if row[id_col].value is None:
            continue
        tid = int(row[id_col].value)
        rec = {}
        for key, idx in cols.items():
            rec[key] = "" if idx is None or row[idx].value is None else str(row[idx].value).strip()
        details[tid] = rec
    return details


def token_grades(token: str) -> tuple[str, frozenset[int]]:
    """Split one raw token into (canonical subject or None, grade set)."""
    if "_" in token:
        base, suffix = token.rsplit("_", 1)
    else:
        base, suffix = token, ""

    sub = c.canonical(base)
    if sub is None:
        return None, frozenset()

    if suffix == "":
        grades = ALL_GRADES
    elif suffix == "S":
        grades = SENIOR_GRADES
    elif suffix == "J":
        grades = JUNIOR_GRADES
    elif suffix.isdigit():
        grades = frozenset({int(suffix)})
    else:
        raise ValueError(f"unknown grade suffix in token {token!r}")
    return sub, grades


def qualifications(teacher: Teacher, existence: dict[str, frozenset[int]]
                   ) -> dict[str, frozenset[int]]:
    """subject -> grades this teacher can teach it in, clamped to existence."""
    result: dict[str, set[int]] = {}
    for token in teacher.tokens:
        sub, grades = token_grades(token)
        if sub is None:
            continue                      # declared non-teaching, ignored
        if sub not in result:
            result[sub] = set()
        result[sub] |= grades             # union across repeated tokens

    # Junior options: grant the option wherever the teacher holds its senior
    # counterpart (constants.JUNIOR_OPTION_SOURCE). Clamping to existence below
    # keeps it honest -- the option is granted only in grades where the roster
    # says it is actually offered.
    for option, source in c.JUNIOR_OPTION_SOURCE.items():
        if source in result:
            result.setdefault(option, set())
            result[option] |= result[source]

    clamped = {}
    for sub, grades in result.items():
        live = grades & existence.get(sub, frozenset())
        if live:
            clamped[sub] = frozenset(live)
    return clamped


def pools(teachers: list[Teacher], existence: dict[str, frozenset[int]]
          ) -> dict[tuple[str, int], frozenset[int]]:
    """(subject, grade) -> set of teacher ids qualified for it."""
    result: dict[tuple[str, int], set[int]] = {}
    for t in teachers:
        for sub, grades in qualifications(t, existence).items():
            for gr in grades:
                result.setdefault((sub, gr), set()).add(t.tid)
    return {k: frozenset(v) for k, v in result.items()}
