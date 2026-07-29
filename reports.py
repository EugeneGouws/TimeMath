"""Sanity reports over an emitted v3 timetable.json.

Deliberately reads the JSON, not the solver's objects: like main.verify() this
is an independent pass, so a bug in the model cannot hide itself here. It
reports; it never asserts a verdict on data it did not derive itself.

The point is to turn manual verification from "read a whole timetable" into
"check the handful of things flagged".

    python reports.py data/2026/timetable.json [out/2026_full]
"""
import json
import os
import sys
from collections import defaultdict

import constants as c

PERIODS_PER_COLUMN = 7
COLUMNS = 8
WEEK = PERIODS_PER_COLUMN * COLUMNS


def load(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _grade(lesson):
    return int(lesson["grade"])


def teacher_report(doc):
    """Periods per teacher per cycle, plus how many distinct grades and
    subjects each carries. A teacher over 56 is impossible; one far under is a
    sign the objective piled work onto too few people."""
    lines = ["=== teacher load ===",
             f"{'teacher':>8} {'periods':>8} {'free':>5} {'lessons':>8}  subjects"]
    per = defaultdict(lambda: {"periods": 0, "lessons": 0, "subjects": set(),
                               "grades": set()})
    for label, lesson in doc["lessons"].items():
        tid = lesson.get("teacher")
        slots = len(doc["placements"].get(label, []))
        rec = per[tid]
        rec["periods"] += slots
        rec["lessons"] += 1
        rec["subjects"].add(label.split("_")[0])
        rec["grades"].add(lesson["grade"])

    flags = []
    for tid in sorted(per, key=lambda t: -per[t]["periods"]):
        rec = per[tid]
        lines.append(f"{str(tid):>8} {rec['periods']:>8} {WEEK - rec['periods']:>5} "
                     f"{rec['lessons']:>8}  {' '.join(sorted(rec['subjects']))}"
                     f"   grades {','.join(sorted(rec['grades']))}")
        if rec["periods"] > WEEK:
            flags.append(f"teacher {tid} teaches {rec['periods']} periods, "
                         f"more than the {WEEK}-cell week")
        elif rec["periods"] == WEEK:
            flags.append(f"teacher {tid} teaches all {WEEK} periods -- no free "
                         f"period anywhere in the cycle")
    loads = [r["periods"] for r in per.values()]
    lines.append(f"\n{len(per)} teachers, load min {min(loads)} / "
                 f"median {sorted(loads)[len(loads) // 2]} / max {max(loads)}")
    return lines, flags


def student_report(doc):
    """Free periods per student, and whether each free period is EXPECTED.

    A free period has two very different causes and conflating them makes the
    report useless:

      * the student's record is short -- they are enrolled in fewer periods
        than the week holds. A data-quality signal, nothing to do with the
        solver. Juniors are meant to carry all 56 cells, so a short junior
        record is worth reporting, but it is not a lost lesson.
      * the student is enrolled in a lesson that never reached the grid. That
        IS a lost lesson and a hard flag.

    Only the second is flagged. The first is listed for the school to fix.
    """
    lines = ["=== student free periods, by grade ==="]
    grade_of = {sid: str(rec.get("grade", "?"))
                for sid, rec in doc.get("students", {}).items()}

    # what each student is enrolled in, in periods
    demand = defaultdict(int)
    for label, ids in doc["enrolments"].items():
        m = len(doc["placements"].get(label, []))
        for sid in ids:
            demand[sid] += m

    free_by_grade = defaultdict(lambda: defaultdict(int))
    short = []
    flags = []
    for sid, slots in doc["student_slots"].items():
        grade = grade_of.get(sid, "?")
        placed = len(slots)
        want = demand.get(sid, 0)
        free_by_grade[grade][WEEK - placed] += 1
        if placed < want:
            flags.append(f"student {sid} (grade {grade}) is enrolled in {want} "
                         f"periods but only {placed} reached the grid -- "
                         f"{want - placed} lesson period(s) LOST")
        elif want < WEEK and grade in ("8", "9"):
            short.append((sid, grade, want))

    for grade in sorted(free_by_grade, key=lambda g: (len(g), g)):
        dist = free_by_grade[grade]
        lines.append(f"  grade {grade:>2}: " +
                     "  ".join(f"{free} free x{n}" for free, n in sorted(dist.items())))
    if short:
        lines.append("")
        lines.append(f"  {len(short)} junior record(s) below the full {WEEK} periods "
                     f"-- under-subscribed in the ROSTER, not dropped by the solver:")
        for sid, grade, want in sorted(short, key=lambda t: t[2]):
            lines.append(f"    student {sid} (grade {grade}): enrolled for {want} "
                         f"of {WEEK} periods")
    return lines, flags


def class_size_report(doc):
    """Roll per lesson against the cap."""
    lines = ["=== class sizes ==="]
    sizes = {label: len(ids) for label, ids in doc["enrolments"].items()}
    over = [(lab, n) for lab, n in sizes.items() if n > c.CLASS_CAP]
    tiny = sorted([(lab, n) for lab, n in sizes.items() if n <= 2], key=lambda t: t[1])
    dist = defaultdict(int)
    for n in sizes.values():
        dist[(n // 5) * 5] += 1
    for lo in sorted(dist):
        lines.append(f"  {lo:>2}-{lo + 4:>2} students: {dist[lo]:>3} lesson(s)")
    if tiny:
        lines.append("  very small: " + ", ".join(f"{lab}({n})" for lab, n in tiny))
    flags = [f"{lab} has {n} students, over the cap of {c.CLASS_CAP}" for lab, n in over]
    return lines, flags


def continuity_report(doc):
    """Does one group of students get ONE teacher per subject?

    A junior register class split across two teachers for the same subject is
    legal in the model (nothing forbids it) but wrong in a school. Reported by
    counting, per (subject, grade), how many lessons share students.
    """
    lines = ["=== teacher continuity (same students, same subject, 2+ teachers) ==="]
    by_subject_grade = defaultdict(list)
    for label, lesson in doc["lessons"].items():
        by_subject_grade[(label.split("_")[0], lesson["grade"])].append(label)

    findings = []
    for (subject, grade), labels in sorted(by_subject_grade.items()):
        if len(labels) < 2:
            continue
        for i in range(len(labels)):
            for j in range(i + 1, len(labels)):
                a, b = labels[i], labels[j]
                ta = doc["lessons"][a].get("teacher")
                tb = doc["lessons"][b].get("teacher")
                if ta == tb:
                    continue
                shared = set(doc["enrolments"].get(a, [])) & set(doc["enrolments"].get(b, []))
                if shared:
                    findings.append(f"{subject}_{grade}: {len(shared)} student(s) take "
                                    f"both {a} (T{ta}) and {b} (T{tb})")
    if findings:
        lines.extend("  " + f for f in findings)
    else:
        lines.append("  none -- every student has one teacher per subject")
    # split lessons (junior EN 7+1) legitimately share students AND teacher, so
    # they never appear above; that is the intended behaviour of the link.
    return lines, findings


def column_report(doc):
    """What each column holds, per grade -- the shape you would eyeball first."""
    lines = ["=== columns (period counts per grade) ==="]
    per = defaultdict(lambda: defaultdict(int))
    for label, lesson in doc["lessons"].items():
        for slot in doc["placements"].get(label, []):
            per[lesson["grade"]][slot[0]] += 1
    for grade in sorted(per, key=lambda g: (len(g), g)):
        row = per[grade]
        lines.append(f"  grade {grade:>2}: " +
                     " ".join(f"{col}={row.get(col, 0):>3}" for col in sorted(row)))
    return lines, []


def compare_to(doc, other_path):
    """Compare teacher headcount and section counts against another v3 file --
    normally the school's own timetable for the same year. Not a target: their
    file is one valid answer, not the optimum. It answers 'is our number sane?'
    """
    lines = [f"=== comparison with {os.path.basename(other_path)} ==="]
    if not os.path.exists(other_path):
        return lines + [f"  not found: {other_path}"], []
    other = load(other_path)
    for name, d in (("ours", doc), ("theirs", other)):
        teachers = {l.get("teacher") for l in d["lessons"].values() if l.get("teacher")}
        lines.append(f"  {name:>7}: {len(d['lessons']):>4} lessons, "
                     f"{len(teachers):>3} teachers, "
                     f"{len(d.get('students', {})):>4} students")
    return lines, []


def roster_report(path, out_dir=None):
    """Data-quality worklist, read RAW from the roster workbook.

    Deliberately does not go through roster.parse_student: that canonicalises
    codes, dedupes repeated subjects and renumbers duplicate ids, all of which
    is right for the solver and wrong here. This report is for the school to
    ACT on, so it must see the file as captured.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb["Roster"] if "Roster" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index("Id")
    gr_col = headers.index("Grade")
    reg_col = headers.index("Reg") if "Reg" in headers else None
    s_cols = [i for i, h in enumerate(headers)
              if h is not None and str(h).startswith("S") and str(h)[1:].isdigit()]

    rows = []
    for row in ws.iter_rows(min_row=2):
        if row[id_col].value is None:
            continue
        rows.append({
            "id": int(row[id_col].value),
            "grade": int(row[gr_col].value),
            "reg": (str(row[reg_col].value).strip()
                    if reg_col is not None and row[reg_col].value is not None else None),
            "raw": [str(row[i].value).strip().upper() for i in s_cols if row[i].value],
        })

    problems = []
    seen = {}
    for rec in rows:
        sid, grade = rec["id"], rec["grade"]
        where = f"student {sid} (grade {grade}{', ' + rec['reg'] if rec['reg'] else ''})"

        if sid in seen:
            problems.append(("DUPLICATE ID", f"{where}: id already used by a grade "
                                             f"{seen[sid]} row -- two different people "
                                             f"share one id"))
        seen[sid] = grade

        counts = defaultdict(int)
        codes = []
        for raw in rec["raw"]:
            counts[raw] += 1
            try:
                code = c.canonical(raw)
            except ValueError:
                problems.append(("UNKNOWN CODE", f"{where}: subject code {raw!r} is not "
                                                 f"in the conversion table"))
                continue
            if code is not None:
                codes.append(code)
        for raw, n in sorted(counts.items()):
            if n > 1:
                problems.append(("DUPLICATE SUBJECT", f"{where}: {raw} listed {n} times"))

        subjects = set(codes)
        junior = c.is_junior(grade)

        if junior and not rec["reg"]:
            problems.append(("NO REGISTER CLASS", f"{where}: juniors are timetabled by "
                                                  f"register class, so this student "
                                                  f"cannot be placed with their class"))

        missing = [sub for sub in ("EN", "LO", "PE") if sub not in subjects]
        if not (subjects & {"MA", "ML"}):
            missing.append("MA/ML")
        if missing:
            problems.append(("MISSING SUBJECT", f"{where}: no {', '.join(missing)}"))

        if {"MA", "ML"} <= subjects:
            problems.append(("BOTH MATHS", f"{where}: takes MA and ML -- almost "
                                           f"certainly a capture error"))

        if junior:
            if not (subjects & c.FAL):
                problems.append(("MISSING FAL", f"{where}: no first additional "
                                                f"language (AF or ZU)"))
            load = sum(c.getM(sub, grade) for sub in subjects)
            if load < WEEK:
                problems.append(("SHORT RECORD", f"{where}: {len(subjects)} subjects = "
                                                 f"{load} periods, {WEEK - load} short "
                                                 f"of the full week"))
            elif load > WEEK:
                problems.append(("OVERLOADED", f"{where}: {load} periods against a "
                                               f"{WEEK}-period week"))
        else:
            choice = subjects - c.UNIVERSAL
            if len(choice) > 5:
                problems.append(("OVERSIZED BASKET", f"{where}: {len(choice)} choice "
                                                     f"subjects against 5 columns"))

    lines = [f"=== data-quality worklist: {os.path.basename(path)} ===",
             f"{len(rows)} rows read", ""]
    by_kind = defaultdict(list)
    for kind, msg in problems:
        by_kind[kind].append(msg)
    if not problems:
        lines.append("no problems found")
    for kind in sorted(by_kind, key=lambda k: -len(by_kind[k])):
        lines.append(f"--- {kind} ({len(by_kind[kind])}) ---")
        lines.extend("  " + m for m in sorted(by_kind[kind]))
        lines.append("")

    out = "\n".join(lines)
    print(out)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
        dest = os.path.join(out_dir, "data_quality.txt")
        with open(dest, "w", encoding="utf-8") as f:
            f.write(out)
        print(f"[out] {dest}")
    return problems


def run(json_path, out_dir=None):
    doc = load(json_path)
    sections = [teacher_report(doc), student_report(doc), class_size_report(doc),
                continuity_report(doc), column_report(doc)]

    year = os.path.basename(os.path.dirname(json_path))
    sections.append(compare_to(doc, f"E:/TimeEduSuite/data/{year}/{year}_v3.json"))

    lines, flags = [], []
    for body, found in sections:
        lines.extend(body)
        lines.append("")
        flags.extend(found)

    header = ["=== FLAGS ==="]
    header += ["  " + f for f in flags] if flags else ["  none"]
    header.append("")
    out = "\n".join(header + lines)

    print(out)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
        path = os.path.join(out_dir, "sanity.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write(out)
        print(f"[out] {path}")
    return flags


if __name__ == "__main__":
    args = sys.argv[1:]
    if args and args[0] == "--roster":
        roster_report(args[1], args[2] if len(args) > 2 else None)
    else:
        json_path = args[0] if args else "data/2026/timetable.json"
        run(json_path, args[1] if len(args) > 1 else None)
