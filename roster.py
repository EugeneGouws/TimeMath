from dataclasses import dataclass
from openpyxl import load_workbook
import constants as c
import math

@dataclass(frozen=True)
class Student:
    id: int
    grade: int
    subjects: frozenset[str]
    reg: str | None = None      # register class, e.g. "09W". Juniors need it.

# functions------------------------------------
def parse_student(path: str) -> list[Student]:
    wb = load_workbook(path, data_only=True)
    # Phase01_<year>.xlsx keeps the roster on a "Roster" sheet; the raw
    # Students<year>.xlsx (the only file carrying juniors and the Reg column)
    # keeps it on the first sheet.
    ws = wb["Roster"] if "Roster" in wb.sheetnames else wb[wb.sheetnames[0]]

    students: list[Student] = []
    header_row = ws[1]
    headers = []
    for cell in header_row:
        headers.append(cell.value)

    id_col = headers.index("Id")
    gr_col = headers.index("Grade")
    reg_col = headers.index("Reg") if "Reg" in headers else None

    # Subject columns are S1..S13 -- match "S" + digits only, so Students<year>
    # .xlsx's other columns are not swept in. Names (Lastname/Firstname) are
    # never read: they must not enter the solver (CLAUDE.md PII policy).
    s_cols = []
    for i in range(len(headers)):
        h = headers[i]
        if h is None:
            continue
        h = str(h)
        if h.startswith("S") and h[1:].isdigit():
            s_cols.append(i)

    for row in ws.iter_rows(min_row=2):
        if row[id_col].value is None:
            continue
        s_id = int(row[id_col].value)
        s_gr = int(row[gr_col].value)
        s_subjects = []

        for i in s_cols:
            sub = row[i].value
            if sub is None:
                continue
            # Canonicalise here, not downstream: NS->SC, CA->CAT, OD->ODR are
            # real drift, and a declared non-teaching code returns None. An
            # unknown code raises -- never a silent default.
            code = c.canonical(str(sub))
            if code is None:
                continue
            if code in s_subjects:
                # Was a hard raise. The raw Students<year>.xlsx has one such
                # row (2026 id 72, LS twice) and it is plainly a capture slip:
                # the student takes the subject once. Refusing to parse the
                # whole school over it helps nobody, so it is deduped LOUDLY.
                print(f"[roster] DATA PROBLEM: student {s_id} lists {code} twice "
                      f"-- deduped, taken once")
                continue
            s_subjects.append(code)

        reg = None
        if reg_col is not None and row[reg_col].value is not None:
            reg = str(row[reg_col].value).strip()

        student = Student(id=s_id, grade=s_gr,
                          subjects=frozenset(s_subjects), reg=reg)
        students.append(student)

    return students

#------------------------------------------------------------
def resolve_duplicate_ids(students: list[Student]) -> tuple[list[Student], list[str]]:
    """Give repeated student Ids fresh synthetic ones, and say so loudly.

    CLAUDE.md correction 5: several Grade 10 rows in the 2026 workbook share
    one Id. Everything downstream keys students by id -- enrol, the y vars,
    the per-student AllDifferent -- so a shared id silently fuses those rows
    into one phantom student holding the union of their baskets (16 subjects,
    two EN sections, both MA and ML). That is exactly PLAN.md §8's plausible
    wrong answer, so it is resolved here, before any modelling.

    They are separate people, so they are renumbered, not merged and not
    dropped: merging invents a basket nobody takes, dropping means those
    students silently receive no timetable.
    """
    seen = set()
    next_id = max(s.id for s in students) + 1
    fixed = []
    report = []

    for s in students:
        if s.id not in seen:
            seen.add(s.id)
            fixed.append(s)
            continue
        report.append(f"duplicate Id {s.id} (grade {s.grade}) renumbered to {next_id}")
        fixed.append(Student(id=next_id, grade=s.grade, subjects=s.subjects))
        seen.add(next_id)
        next_id += 1

    return fixed, report

#------------------------------------------------------------
def quarantine(students: list[Student], choice_cols: int = 5,
               band_cells: int = 14) -> tuple[list[Student], list[Student], list[str]]:
    """Apply the exception rule (CLAUDE.md) before any modelling.

    Only OVERSIZED records affect feasibility -- a student who needs more
    columns, or more band periods, than the grid physically has. An undersized
    basket is a subset constraint and constrains nothing, so it is reported as
    a data-quality signal and kept in the model.

    Returns (kept, quarantined, report). Quarantined students must be
    reinstated at Phase 5 output or they silently receive no timetable.
    """
    kept, out, report = [], [], []

    for s in students:
        # --- juniors: different rule, because there is no band and no basket.
        # A junior takes essentially the whole timetable (55 of 56 cells on
        # 2026 data), so the only thing that can make them impossible is a
        # total period load the week cannot hold.
        if c.is_junior(s.grade):
            total = 0
            for sub in s.subjects:
                total += c.getM(sub, s.grade)
            if total > 56:
                out.append(s)
                report.append(f"student {s.id} (grade {s.grade}) OVERLOADED: "
                              f"{len(s.subjects)} subjects = {total} periods but the "
                              f"week holds 56 -- quarantined")
            else:
                kept.append(s)
                if s.reg is None:
                    report.append(f"student {s.id} (grade {s.grade}) has NO register "
                                  f"class -- junior sections are built from register "
                                  f"classes, so this student cannot be placed with "
                                  f"their class")
            continue

        choice = s.subjects - c.UNIVERSAL
        band_load = 0
        for sub in s.subjects & c.BAND:
            band_load += c.getM(sub, s.grade)

        if len(choice) > choice_cols:
            out.append(s)
            report.append(f"student {s.id} (grade {s.grade}) OVERSIZED BASKET: "
                          f"{len(choice)} choice subjects {sorted(choice)} but only "
                          f"{choice_cols} choice columns -- quarantined, needs a "
                          f"subject moved out-of-timetable (P1-P4)")
        elif band_load > band_cells:
            out.append(s)
            # Two different causes, and conflating them hides the second one:
            # MA+ML together is a capture error, but a junior grade overflows
            # honestly (MA 10 + LO 3 + PE 2 = 15), because getM's junior
            # override makes LO 3 -- that needs a WIDER band, not a quarantine.
            if {"MA", "ML"} <= s.subjects:
                why = "MA and ML together is almost certainly a capture error"
            else:
                why = (f"band width is hardcoded at 2 columns ({band_cells} cells); "
                       f"this record needs ceil({band_load}/7) columns")
            report.append(f"student {s.id} (grade {s.grade}) OVERSIZED BAND: "
                          f"{sorted(s.subjects & c.BAND)} = {band_load} periods but "
                          f"the band holds {band_cells} -- quarantined ({why})")
        else:
            kept.append(s)
            # An ML student legitimately carries ML(7)+LO(2)+PE(2) = 11 of 14
            # and 3 slack cells, so a short band is NOT by itself incomplete.
            # What is incomplete is a MISSING band subject.
            missing = []
            if not (s.subjects & {"MA", "ML"}):
                missing.append("MA/ML")
            for sub in ("LO", "PE", "EN"):
                if sub not in s.subjects:
                    missing.append(sub)
            if missing or not choice:
                report.append(f"student {s.id} (grade {s.grade}) INCOMPLETE RECORD: "
                              f"missing {missing or 'nothing'}, {len(choice)} choice "
                              f"subjects -- kept in the model, constrains nothing")

    return kept, out, report

#------------------------------------------------------------
def existence(students: list[Student]) -> dict[str, frozenset[int]]:
    table = {}
    for student in students:
        for s in student.subjects:
            if s not in table:
                table[s] = set()
            table[s].add(student.grade)

    result = {}
    for s in table:
        result[s] = frozenset(table[s])
    return result

#---------------------------------------------------------------
def enrol(students: list[Student], gr: int) -> dict[str, frozenset[int]]:
    table = {}
    for student in students:
        if student.grade != gr:
            continue
        for s in student.subjects:
            if s not in table:
                table[s] = set()
            table[s].add(student.id)

    result = {}
    for s in table:
        result[s] = frozenset(table[s])
    return result

#-----------------------------------------------------------
def floors(students: list[Student], gr: int, cap: int = c.CLASS_CAP) -> dict[str, int]:
    e = enrol(students, gr)

    result = {}
    for s in e:
        result[s] = math.ceil(len(e[s]) / cap)
    return result

#----------------------------------------------------------
def baskets(students: list[Student], gr: int) -> dict[frozenset[str], frozenset[int]]:
    table = {}
    for student in students:
        if student.grade != gr:
            continue
        key = student.subjects - c.UNIVERSAL
        if key not in table:
            table[key] = set()
        table[key].add(student.id)

    result = {}
    for s in table:
        result[s] = frozenset(table[s])
    return result
